from openpyxl import load_workbook
import xml.etree.ElementTree as ET
from datetime import datetime
import json,os,pytz, uuid,zipfile
import tkinter as tk
import tkinter.messagebox as messagebox
from openpyxl.utils.exceptions import SheetTitleException

def load_config(filename:str):
    with open(filename, "r", encoding='utf-8') as configfile:
        config_json = configfile.read()
    config = json.loads(config_json)
    return config

def current_datetime(value=None):
    tz = pytz.timezone("Europe/Moscow")
    if value is None:
        now = datetime.now(tz).replace(microsecond=0)
    else:
        if isinstance(value, datetime):
            if value.tzinfo is None or value.tzinfo.utcoffset(value) is None:
                now = tz.localize(value.replace(microsecond=0), is_dst=None)
            else:
                now = value.astimezone(tz).replace(microsecond=0)
        elif isinstance(value, str):
            value = value.replace('г.', '').strip()
            try:
                parsed_date = datetime.strptime(value, "%d.%m.%Y")
                now = tz.localize(parsed_date.replace(hour=0, minute=0, second=0, microsecond=0), is_dst=None)
            except ValueError as e:
                messagebox.showerror("Ошибка", f"Ошибка преобразования даты: {e}")
                return None
        else:
            messagebox.showwarning("Внимание", "Недопустимое значение даты в ячейке")
            return None
    now_moscow_iso = now.isoformat().replace('+03:00', 'Z')
    return now_moscow_iso

def load_workbook_data(filename):
    wb = load_workbook(filename=filename, read_only=True)
    return wb




def add_to_zip(filepath,INN,UID):
    zip_name = os.path.join(os.path.dirname(filepath), f"{INN}_{datetime.now().strftime('%Y%m%d')}_{UID}.zip")
    print(zip_name)
    with zipfile.ZipFile(zip_name,'w') as zipf:
       zipf.write(filepath,os.path.basename(filepath))
       return filepath 
 
def find_first_row_with_prefix(sheet, start_row, prefix):
    for row in range(start_row, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=1).value
        if isinstance(cell_value, str) and cell_value.startswith(prefix):
            return row
    return None

def create_xml(sheet1,sheet2, config, now_local):
    thisUID = str(uuid.uuid4())
    reportdate=str(now_local)
    root =ET.Element("Message")
    root.set('CreateDate',str(now_local))
    root.set('UID',str(thisUID))
    root.set('PreviousUID',str(config["PreviousUID"]))
    config["PreviousUID"] = thisUID
    Organization =ET.SubElement(root, "Organization")
    Organization.set("INN",str(config["INN"]))
    Organization.set("KPP",config["KPP"])
    Organization.set('Name',config["NAME"])
    Organization.set("GOZUID",str(sheet1[config["GOZUID"]].value))
    Organization.set("ContractDate",str(current_datetime(sheet1[config["start_indexes"]["ContractDate"]].value))) 
                                                 
    Forms=ET.SubElement(root,"Forms")
    Form8=ET.SubElement(Forms,"Form8",{"ReportDate":reportdate,'Year':str(sheet1[config['start_indexes']['Year']].value),
                                      'Quarter':str(sheet1[config["start_indexes"]["Quarter"]].value)})
    ContractSpending=ET.SubElement(Form8,"ContractSpending")
    Contractors = ET.SubElement(Form8, 'Contractors')
    PlannedPay=ET.SubElement(Form8,"PlannedPay")
    ContractFinance=ET.SubElement(Form8,"ContractFinance")
    for row in range(7, sheet1.max_row + 1):
      if str(sheet1.cell(row=row, column=1).value)=='1':  
        ContractSpending.set("Total",str(int(sheet1.cell(row=row, column=3).value*100)))
      if str(sheet1.cell(row=row, column=1).value).startswith('1.1'):  
        ContractSpending.set("Salary",str(int(sheet1.cell(row=row, column=3).value*100)))
      if str(sheet1.cell(row=row, column=1).value).startswith('1.2'):  
        ContractSpending.set("Taxes",str(int(sheet1.cell(row=row, column=3).value*100)))
      if str(sheet1.cell(row=row, column=1).value).startswith('1.3'):   
        ContractSpending.set("Rates",str(int(sheet1.cell(row=row, column=3).value*100)))
      if str(sheet1.cell(row=row, column=1).value).startswith('1.4'):    
        ContractSpending.set("Other",str(int(sheet1.cell(row=row, column=3).value*100)))
      if str(sheet1.cell(row=row, column=1).value).startswith('1.5'):  
        ContractSpending.set("Reserve",str(int(sheet1.cell(row=row, column=3).value*100)))
      if str(sheet1.cell(row=row, column=1).value).startswith('1.6'):  
        ContractSpending.set("Income",str(int(sheet1.cell(row=row, column=3).value*100)))
      if str(sheet1.cell(row=row, column=1).value).startswith('2.'):
       Contractor = ET.SubElement(Contractors,"Contractor",{'Total':str(int(sheet1.cell(row=row, column=3).value*100)),
                                                           'Name':str(sheet1.cell(row=row, column=4).value),
                                                           'INN':str(sheet1.cell(row=row, column=5).value),
                                                           'ContractNumber':str(sheet1.cell(row=row, column=6).value),
                                                           'ContractDate': str(current_datetime(sheet1.cell(row=row, column=7).value)),
                                                           'AccountNumber':str(sheet1.cell(row=row, column=8).value),
                                                           'Cost': str(int(sheet1.cell(row=row, column=9).value*100)),
                                                           'PaymentPlanned':str(int(sheet1.cell(row=row, column=10).value*100)),
                                                           'PaymentCurrent':str(int(sheet1.cell(row=row, column=11).value*100)),
                                                           'FinishDate': str(current_datetime(sheet1.cell(row=row, column=12).value))})
      if str(sheet1.cell(row=row, column=1).value)=='3':
       PlannedPay.set("Total", str(int(sheet1.cell(row=row, column=9).value*100))) #цена договора?
       PlannedPay.set("PaymentPlanned", str(int(sheet1.cell(row=row, column=10).value*100)))
       PlannedPay.set("PaymentCurrent", str(int(sheet1.cell(row=row, column=3).value*100)))
     
      if str(sheet1.cell(row=row, column=1).value)=='4':
       ContractFinance.set("TotalRequirement",str(int(sheet1.cell(row=row, column=3).value*100)))
      if str(sheet1.cell(row=row, column=1).value)=='5': 
       ContractFinance.set("CashBalance",str(int(sheet1.cell(row=row, column=3).value*100)))
      ContractFinance.set("DateBalance",reportdate)
      if str(sheet1.cell(row=row, column=1).value)=='6':
       ContractFinance.set("PlannedIncome",str(int(sheet1.cell(row=row, column=3).value*100)))
      if str(sheet1.cell(row=row, column=1).value)=='7':
       ContractFinance.set("DepositeIncome",str(int(sheet1.cell(row=row, column=3).value*100)))
    
    FormSupplement=ET.SubElement(Forms,'Supplement',{'ReportDate':reportdate})
    Parts=ET.SubElement(FormSupplement,"Parts")
    for row in range(3, sheet1.max_row + 1):
      if str(sheet2.cell(row=row, column=1).value).startswith('202'):
        Part=ET.SubElement(Parts,"Part",{"Year":str(sheet2.cell(row=row, column=1).value),
                                  "Quarter":str(sheet2.cell(row=row, column=2).value),
                                  "Requirement":str(int(sheet2.cell(row=row, column=3).value*100)),
                                  "Deviation":str(int(sheet2.cell(row=row, column=4).value*100))})
       
        reasons = str(sheet2.cell(row=row, column=5).value).split(',')
        
        Reasons = ET.SubElement(Part,"Reasons")
        for reason in reasons:
         if  reason.strip() != 'None': #and reason.strip() != '':
            ET.SubElement(Reasons,"Reason").text=reason.strip()
       
    from xml.dom.minidom import parseString  
    xmlstr = parseString(ET.tostring(root,encoding="UTF-8", xml_declaration=True)).toprettyxml(encoding="UTF-8")
    
    return xmlstr,config

def update_config(filename, config):
    with open(filename, "w", encoding='utf-8') as configfile:
        config_json = json.dumps(config, ensure_ascii=False)
        configfile.write(config_json)

def save_xml(xmlstr, filename):
    with open(filename, "wb") as f:
        f.write(xmlstr)

def main():
    root = tk.Tk()
    root.withdraw()
    
    from tkinter import filedialog
    xlsx_file_path = filedialog.askopenfilename(title="Выберите XLSX файл", filetypes=[("XLSX Files", "*.xlsx")])
    config = load_config("config.json")
    try:
     workbook = load_workbook_data(xlsx_file_path)
     sheet1 = workbook.worksheets[0]
     sheet2 = workbook.worksheets[1]
     xmlstr,updated_config = create_xml(sheet1,sheet2, config, current_datetime())
    except SheetTitleException as e:
       messagebox.ERROR(e)
    save_xml(xmlstr, os.path.dirname(xlsx_file_path) + "\message.xml")
    update_config("config.json", updated_config) 
    add_to_zip(os.path.dirname(xlsx_file_path)+'\message.xml',config["INN"],config["PreviousUID"])
    messagebox.showinfo("SAIBIS","XML отчет сохранен рядом с " + xlsx_file_path)
   
    from lxml import etree
    try:
      with open('message_v4.xsd', 'rb') as f:
       schema_root = etree.XML(f.read())
      schema = etree.XMLSchema(schema_root)
      try:     
       with open(os.path.dirname(xlsx_file_path)+'\message.xml', 'rb') as f:
        xml_root = etree.XML(f.read())
       try:
        schema.assertValid(xml_root)
        messagebox.showinfo("saibis","XML соответствует XSD схеме.")
       except etree.DocumentInvalid as err:
        messagebox.showinfo( "Ошипка","Ошибка валидации:"+ str(err))
      except IOError as err:
       print(err) 
    except IOError as err:
         print(err)
    
main()
