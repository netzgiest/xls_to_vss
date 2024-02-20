from openpyxl import load_workbook
import xml.etree.ElementTree as ET
from datetime import datetime
import json,os,pytz, uuid,zipfile
import tkinter as tk
import tkinter.messagebox as messagebox
from openpyxl.utils.exceptions import SheetTitleException

def validate_xml(xml_str, schema_path):
    from lxml import etree
    from pathlib import Path
    schema_root = etree.fromstring(Path(schema_path).read_bytes())
    schema = etree.XMLSchema(schema_root)
    xml_root = etree.fromstring(xml_str)
    try:
        schema.assertValid(xml_root)
        return "XML соответствует XSD схеме."
    except etree.DocumentInvalid as err:
        return f"Ошибка валидации: {err}"

def load_config(filename:str):
    with open(filename, "r", encoding='utf-8') as configfile:
        config_json = configfile.read()
    config = json.loads(config_json)
    return config
    
def current_datetime(value=None):
    tz = pytz.timezone("Europe/Moscow")
    if value is None:
        now = datetime.now(tz).replace(microsecond=0)
    else:
        if isinstance(value, datetime):
            # Если value уже имеет установленный часовой пояс, использовать его
            # Иначе считаем что время передано в часовом поясе Москвы
            if value.tzinfo is None or value.tzinfo.utcoffset(value) is None:
                now = tz.localize(value.replace(microsecond=0), is_dst=None)
            else:
                now = value.astimezone(tz).replace(microsecond=0)
        elif isinstance(value, str):
            value = value.replace('г.', '').strip()
            try:
                # Преобразовываем строку в дату
                parsed_date = datetime.strptime(value, "%d.%m.%Y")
                # Локализуем дату для часового пояса Москвы
                now = tz.localize(parsed_date.replace(hour=0, minute=0, second=0, microsecond=0), is_dst=None)
            except ValueError as e:
                messagebox.showerror("Ошибка", f"Ошибка преобразования даты: {e}")
                return None
        else:
            messagebox.showwarning("Внимание", "Недопустимое значение даты в ячейке")
            return None
    # Форматируем время в ISO строки и заменяем '+03:00' на 'Z', чтобы указать время по Москве
    now_moscow_iso = now.isoformat().replace('+03:00', 'Z')
    return now_moscow_iso

def load_workbook_data(filename):
    wb = load_workbook(filename=filename, read_only=True)
    return wb

def add_to_zip(filepath,INN,UID):
    zip_name = os.path.dirname(filepath) + f"{INN}_{datetime.now().strftime('%Y%m%d')}_{UID}.zip"
    print(zip_name)
    with zipfile.ZipFile(zip_name,'w') as zipf:
       zipf.write(filepath)
       return filepath 


def add_child_element(parent, tag, sheet, row, column):
    try:
        cell_value = sheet.cell(row=row, column=column).value
        if isinstance(cell_value, (float,str)) or cell_value == 0:
            # Конвертируем значение в копейки и добавляем элемент.
            parent.set(tag, str(int(cell_value*100)))
        elif isinstance(cell_value, str) and cell_value.strip() in ["-", "Х",'None']:
            # Если содержимое ячейки - символ "-" или "Х", сообщаем об этом.
            messagebox.showwarning("Внимание", f"Обнаружен недопустимый символ в ячейке ({row}, {column}): {cell_value}")
        else:
            messagebox.showerror("Ошибка", f"В ячейке ({row}, {column}) неизвестное содержимое: {cell_value}")
    except Exception as e:
        messagebox.showerror("Ошибка", f"Произошла ошибка при обработке ячейки ({row}, {column}): {e}")


def create_xml(sheet1,sheet2, config, now_local):
    thisUID = str(uuid.uuid4())
    reportdate=str(now_local)
    root = create_xml_element("Message", {'CreateDate': now_local, 'UID': thisUID, 'PreviousUID': config['PreviousUID']})
    config["PreviousUID"] = thisUID
    Organization = create_xml_element("Organization", {"INN": config["INN"], "KPP": config["KPP"], "Name": config["NAME"],"GOZUID": sheet1[config["GOZUID"]].value,
        "ContractDate":str(current_datetime(sheet1[config["start_indexes"]["ContractDate"]].value))})
    root.append(Organization)

    Forms = create_xml_element("Forms") 
    Form8 = create_xml_element("Form8", {"ReportDate": str(now_local),"Year": str(sheet1[config['start_indexes']['Year']].value),"Quarter" : str(sheet1[config['start_indexes']["Quarter"]].value)}) 
    root.append(Forms)
    Forms.append(Form8)
    ContractSpending = create_xml_element("ContractSpending")
    Form8.append(ContractSpending)
    for row in range(8, sheet1.max_row + 1):
     cell_value = str(sheet1.cell(row=row, column=1).value).strip()

     indexstr = {'1': 'Total','1.1': 'Salary','1.2': 'Taxes','1.3': 'Rates','1.4': 'Other','1.5': 'Reserve',
        '1.6': 'Income','2': 'Contractors','3': 'PlannedPay','4': 'ContractFinance'}
     tag = indexstr.get(cell_value)
     if tag:
        if tag == 'Contractors':
            Contractors = create_xml_element('Contractors')
            Form8.append(Contractors)
        elif tag == 'PlannedPay':
            PlannedPay = create_xml_element('PlannedPay',{
            "Total": str(int(sheet1.cell(row=row, column=9).value*100)), #цена договора?
            "PaymentPlanned": str(int(sheet1.cell(row=row, column=10).value*100)),
            "PaymentCurrent": str(int(sheet1.cell(row=row, column=3).value*100))})
            # Читаем содержимое ячейки и устанавливаем как текст элемента
            Form8.append(PlannedPay)
        elif tag =='ContractFinance':
           ContractFinance = create_xml_element('ContractFinance',
           {"TotalRequirement":str(int(sheet1.cell(row=row, column=3).value*100)),
           "CashBalance":str(int(sheet1.cell(row=row+1, column=3).value*100)),
           "DateBalance":reportdate,
           "PlannedIncome":str(int(sheet1.cell(row=row+2, column=3).value*100)),
           "DepositeIncome":str(int(sheet1.cell(row=row+3, column=3).value*100))})
           Form8.append(ContractFinance)             
        else:
            add_child_element(ContractSpending, tag, sheet1, row, 3)
     else:
        parent_tag = cell_value.split('.')[0]
        main_tag = indexstr.get(parent_tag, None)
        if main_tag:
            if main_tag == 'Contractors':
                Contractor = create_xml_element('Contractor',{'Total':str(int(sheet1.cell(row=row, column=3).value*100)),
                                                            'Name':str(sheet1.cell(row=row, column=4).value),
                                                            'INN':str(sheet1.cell(row=row, column=5).value),
                                                            'ContractNumber':str(sheet1.cell(row=row, column=6).value),
                                                            'ContractDate': str(current_datetime(sheet1.cell(row=row, column=7).value)),
                                                            'AccountNumber':str(sheet1.cell(row=row, column=8).value),
                                                            'Cost': str(int(sheet1.cell(row=row, column=9).value*100)),
                                                            'PaymentPlanned':str(int(sheet1.cell(row=row, column=10).value*100)),
                                                            'PaymentCurrent':str(int(sheet1.cell(row=row, column=11).value*100)),
                                                            'FinishDate': str(current_datetime(sheet1.cell(row=row, column=12).value))})
                Contractors.append(Contractor)
    
    FormSupplement=create_xml_element('Supplement',{'ReportDate':reportdate})
    Forms.append(FormSupplement)
    Parts=ET.SubElement(FormSupplement,"Parts")
    for row in range(3, sheet1.max_row + 1):
      if str(sheet2.cell(row=row, column=1).value).startswith('202'):
        Part=ET.SubElement(Parts,"Part",{"Year":str(sheet2.cell(row=row, column=1).value),
                                  "Quarter":str(sheet2.cell(row=row, column=2).value),
                                  "Requirement":str(int(sheet2.cell(row=row, column=3).value*100)),
                                  "Deviation":str(int(sheet2.cell(row=row, column=4).value*100))})
       
        reasons = str(sheet2.cell(row=row, column=5).value).split(',')
        Reasons = ET.SubElement(Part,"Reasons")
        for reason in reasons:
         if  reason.strip() != 'None': #and reason.strip() != '':
            ET.SubElement(Reasons,"Reason").text=reason.strip()
    
    from xml.dom.minidom import parseString  
    xmlstr = parseString(ET.tostring(root,encoding='UTF-8', xml_declaration=True)).toprettyxml(encoding="UTF-8")
    
    return xmlstr,config

def create_xml_element(tag, attributes=None, text=None):
    element = ET.Element(tag)
    if attributes:
        for key, value in attributes.items():
            element.set(key, value)
    if text:
        element.text = text
    return element

def update_config(filename, config):
    with open(filename, "w", encoding='utf-8') as configfile:
        config_json = json.dumps(config, ensure_ascii=False)
        configfile.write(config_json)

def save_xml(xmlstr, filename):
    with open(filename, "wb") as f:
        f.write(xmlstr)

def main():
    root = tk.Tk()
    root.withdraw()
    
    from tkinter import filedialog
    xlsx_file_path = 'd:\pole.xlsx' #filedialog.askopenfilename(title="Выберите XLSX файл", filetypes=[("XLSX Files", "*.xlsx")])
    config = load_config("config.json")
    try:
     workbook = load_workbook_data(xlsx_file_path)
     sheet1 = workbook.worksheets[0]
     sheet2 = workbook.worksheets[1]
     xmlstr,updated_config = create_xml(sheet1,sheet2, config, current_datetime())
    except SheetTitleException as e:
       messagebox.ERROR(e)
    save_xml(xmlstr, os.path.dirname(xlsx_file_path) + "\message1.xml")
    update_config("config.json", updated_config) 
   # add_to_zip(os.path.dirname(xlsx_file_path)+'\message.xml',config["INN"],config["PreviousUID"])
   # messagebox.showinfo("SAIBIS","XML отчет сохранен рядом с " + xlsx_file_path)
    validation_message = validate_xml(xmlstr, 'message_v4.xsd')
   # messagebox.showinfo("saibis",validation_message)
    if validation_message.__contains__("Ошибка валидации"):
       messagebox.showinfo("saibis",validation_message)
    else:
        print(validation_message)
    
    
main()
