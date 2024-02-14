from openpyxl import load_workbook
from xml.dom.minidom import parseString
from xml.etree.ElementTree import Element, SubElement, tostring
from datetime import datetime
import json,os,pytz, uuid,zipfile
import tkinter as tk
from tkinter import filedialog
import tkinter.messagebox as messagebox

def load_config(filename):
    with open(filename, "r", encoding='utf-8') as configfile:
        config_json = configfile.read()
    config = json.loads(config_json)
    return config

def create_sub_element(parent, name, value=None):
    element = SubElement(parent, name)
    try:
        element.text = str(value)
    except (TypeError, ValueError) as e:
        print(f"Error in setting text of the element: {e}")
    return element

def current_datetime(value=None):
    tz = pytz.timezone("Europe/Moscow")
    if value is None:
     now_utc = datetime.now(pytz.UTC)
     now_local = now_utc.astimezone(tz)
    else:
     now_local= tz.localize(datetime.strptime(str(value), "%d.%m.%Y")).replace(hour=0, minute=0, second=0)  
    return now_local.isoformat()

def load_workbook_data(filename):
    wb = load_workbook(filename=filename, read_only=True)
    return wb

def prettify(elem):
    rough_string = tostring(elem, 'utf-8')
    reparsed = parseString(rough_string)
    return reparsed.toprettyxml(indent="   ", encoding="UTF-8")

def add_to_zip(filepath,INN,UID):
    zip_name = f"{INN}_{datetime.now().strftime('%Y%m%d')}_{UID}.zip"

    with zipfile.ZipFile(zip_name,'w') as zipf:
       zipf.write(filepath)
       return filepath 

def create_xml(sheet1,sheet2, config, now_local):
    root = Element('Message')
    create_sub_element(root, "CreateDate",str(now_local))
    thisUID = str(uuid.uuid4())
    create_sub_element(root, "UID",thisUID)                       
    create_sub_element(root, "PreviousUID",config["PreviousUID"])
    config["PreviousUID"] = thisUID
    Organization = SubElement(root, "Organization")
    create_sub_element(Organization, "INN",config["INN"])
    create_sub_element(Organization,"KPP",config["KPP"])
    create_sub_element(Organization,"Name",config["NAME"])
    create_sub_element(Organization,"GOZUID",str(sheet1[config["start_indexes"]["GOZUID"]].value))
    create_sub_element(Organization,"ContractDate",str(current_datetime(sheet1[config["start_indexes"]["ContractDate"]].value)))
    Forms = SubElement(root,"Forms")
    Form8=SubElement(Forms,"Form",{'type':'8'})
    create_sub_element(Form8,"ReportDate",str(now_local))
    create_sub_element(Form8,"Year",str(sheet1[config["start_indexes"]["Year"]].value))
    create_sub_element(Form8,"Quater",str(sheet1[config["start_indexes"]["Ouarter"]].value))
    Form8ContractSending=SubElement(Form8,"ContractSpending")
    create_sub_element(Form8ContractSending,"Salary",str(sheet1[config["start_indexes"]["Salary"]].value*100))
    create_sub_element(Form8ContractSending,"Taxes",str(sheet1[config["start_indexes"]["Taxes"]].value*100))
    create_sub_element(Form8ContractSending,"Rates",str(sheet1[config["start_indexes"]["Rates"]].value*100))
    create_sub_element(Form8ContractSending,"Other",(sheet1[config["start_indexes"]["Other"]].value*100))
    create_sub_element(Form8ContractSending,"Reserve",str(sheet1[config["start_indexes"]["Reserve"]].value*100))
    create_sub_element(Form8ContractSending,"Income",str(int(sheet1[config["start_indexes"]["Income"]].value)*100))
    contractors = SubElement(Form8, 'Contractors') 
    nextCell = int(config["start_indexes"]["startcontractor"])
    for row in range(nextCell, sheet1.max_row + 1):
     if str(sheet1.cell(row=row, column=1).value).startswith('2'):
      contractor = SubElement(contractors,"Contractor") 
      create_sub_element(contractor, 'Total',str(int(sheet1.cell(row=row, column=3).value)*100))
      create_sub_element(contractor, 'Name',str(sheet1.cell(row=row, column=4).value))
      create_sub_element(contractor, 'INN',str(sheet1.cell(row=row, column=5).value))
      create_sub_element(contractor, 'ContractNumber',str(sheet1.cell(row=row, column=6).value))
      create_sub_element(contractor, 'ContractDate', str(current_datetime(sheet1.cell(row=row, column=7).value)))
      create_sub_element(contractor, 'AccountNumber',str(sheet1.cell(row=row, column=8).value))
      create_sub_element(contractor, 'Cost', str(int(sheet1.cell(row=row, column=9).value)*100))
      create_sub_element(contractor, 'PaymentPlanned',str(int(sheet1.cell(row=row, column=10).value)*100))
      create_sub_element(contractor, 'PaymentCurrent', str(int(sheet1.cell(row=row, column=11).value)*100))
      create_sub_element(contractor, 'FinishDate', str(current_datetime(sheet1.cell(row=row, column=12).value)))    
      nextCell=row+1
    PlannedPay=SubElement(Form8,"PlannedPay")
    create_sub_element(PlannedPay,"PaymentPlanned",str(int(sheet1['J'+str(nextCell)].value)*100))
    create_sub_element(PlannedPay,"Total",str(int(sheet1['I'+str(nextCell)].value)*100)  )
    create_sub_element(PlannedPay,"PlannedPay",str(int(sheet1['C'+str(nextCell)].value)*100))
    ContractFinance=SubElement(Form8,"ContractFinance")
    create_sub_element(ContractFinance,"TotalRequirement",str(int(sheet1['C'+str(nextCell+1)].value)*100))
    create_sub_element(ContractFinance,"CashBalance",str(int(sheet1['C'+str(nextCell+2)].value)*100))
    create_sub_element(ContractFinance,"PlannedIncome",str(int(sheet1['C'+str(nextCell+3)].value)*100))
    create_sub_element(ContractFinance,"DepositeIncome",str(int(sheet1['C'+str(nextCell+4)].value)*100))
   
    FormSupplement=SubElement(Forms,"Form",{'type':'Supplement'})
    create_sub_element(FormSupplement, "ReportDate",str(now_local))
    Parts=SubElement(FormSupplement,"Parts")
   # хз будем считать что с 3 строки и год начинается с 202 и данные начинаются с 1 ячейки
    for row in range(3, sheet1.max_row + 1):
      if str(sheet2.cell(row=row, column=1).value).startswith('202'):
        create_sub_element(Parts,"Year",str(sheet2.cell(row=row, column=1).value))
        create_sub_element(Parts,"Quarter",str(sheet2.cell(row=row, column=2).value))
        create_sub_element(Parts,"Requirement",str(int(sheet2.cell(row=row, column=3).value)*100))
        create_sub_element(Parts,"Deyiation",str(int(sheet2.cell(row=row, column=4).value)*100))
        Reasons = SubElement(Parts,"Reasons")
        reasons = str(sheet2.cell(row=row, column=5).value).split(',')
        for reason in reasons:
            create_sub_element(Reasons,"Reason", reason.strip())
    xmlstr = prettify(root)
    return xmlstr,config

def update_config(filename, config):
    with open(filename, "w", encoding='utf-8') as configfile:
        config_json = json.dumps(config, ensure_ascii=False)
        configfile.write(config_json)

def save_xml(xmlstr, filename):
    with open(filename, "wb") as f:
        f.write(xmlstr)

def main():
    root = tk.Tk()
    root.withdraw()
    xlsx_file_path = filedialog.askopenfilename(title="Выберите XLSX файл", filetypes=[("XLSX Files", "*.xlsx")])
    config = load_config("config.json")
    workbook = load_workbook_data(xlsx_file_path)
    sheet1 = workbook.worksheets[0]
    sheet2 = workbook.worksheets[1]
    xmlstr,updated_config = create_xml(sheet1,sheet2, config, current_datetime())
    save_xml(xmlstr, os.path.dirname(xlsx_file_path) + "/message.xml")
    update_config("config.json", updated_config) 
    add_to_zip(os.path.dirname(xlsx_file_path)+'\message.xml',config["INN"],config["PreviousUID"])
    messagebox.showinfo("SAIBIS","XML отчет сохранен рядом с " + xlsx_file_path)
    
main()