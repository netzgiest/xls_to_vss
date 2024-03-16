from openpyxl import load_workbook
import xml.etree.ElementTree as ET
from datetime import datetime
import json, os, pytz, uuid, zipfile, re
import tkinter as tk
import tkinter.messagebox as messagebox
from openpyxl.utils.exceptions import SheetTitleException


def load_config(filename: str):
    with open(filename, "r", encoding='utf-8') as configfile:
        config_json = configfile.read()
    config = json.loads(config_json)
    return config


def current_datetime(value=None):
    tz = pytz.timezone("Europe/Moscow")
    if value is None:
        now = datetime.now(tz).replace(microsecond=0)
    else:
        if isinstance(value, datetime):
            if value.tzinfo is None or value.tzinfo.utcoffset(value) is None:
                now = tz.localize(value.replace(microsecond=0), is_dst=None)
            else:
                now = value.astimezone(tz).replace(microsecond=0)
        elif isinstance(value, str):
            value = value.replace('г.', '').strip()
            try:
                parsed_date = datetime.strptime(value, "%d.%m.%Y")
                now = tz.localize(parsed_date.replace(hour=0, minute=0, second=0, microsecond=0), is_dst=None)
            except ValueError as e:
                messagebox.showerror("Ошибка", f"Ошибка преобразования даты: {e}")
                return None
        else:
            messagebox.showwarning("Внимание", "Недопустимое значение даты в ячейке")
            return None
    now_moscow_iso = now.isoformat().replace('+03:00', 'Z')
    return now_moscow_iso

def parse_quarter_year(string_to_parse):
    pattern = re.compile(r'(\b[IVXLCDM]+\b|\b\d+\b)(?:.*?)(\b\d{4}\b)')
    match = pattern.search(string_to_parse)
    if match:
        roman_to_arabic = {
            'I': 1, 'II': 2, 'III': 3, 'IV': 4, 'V': 5, 'VI': 6,
            'VII': 7, 'VIII': 8, 'IX': 9, 'X': 10, 'XI': 11, 'XII': 12
        }
        quarter_str = match.group(1).upper()
        if quarter_str in roman_to_arabic:
            quarter = roman_to_arabic[quarter_str]
        else:
           quarter = quarter_str
        year = match.group(2)
        return str(quarter), str(year)
    else:
        raise ValueError("Строка не содержит информацию о квартале и(или) годе в требуемом формате.")

def parse_contract_info(string_to_parse):
    date_pattern = re.compile(r'\b\d{2}\.\d{2}\.(20\d{2})\b')
    id_pattern = re.compile(r'\b\d{25}\b')
    year_match = date_pattern.search(string_to_parse)
    id_match = id_pattern.search(string_to_parse)
    
    if year_match and id_match:
       year = year_match.group(0)  
       contract_id = id_match.group(0)
       return year, contract_id
    else:
        raise ValueError("Строка не содержит информации о годе или идентификаторе контракта в требуемом формате.")
    
def parse_date_creation(string_to_parse):
    date_pattern = re.compile(r'\b\d{2}\.\d{2}\.(20\d{2})\b')
    year_match = date_pattern.search(string_to_parse)
    if year_match:
       year = year_match.group(0)  
       return year
    else:
        raise ValueError("Строка не содержит информации о годе или идентификаторе контракта в требуемом формате.")
def load_workbook_data(filename):
    wb = load_workbook(filename=filename, read_only=True)
    return wb

def add_to_zip(filepath, INN, UID):
    zip_name = os.path.join(os.path.dirname(filepath), f"{INN}_{datetime.now().strftime('%Y%m%d')}_{UID}.zip")
    print(zip_name)
    with zipfile.ZipFile(zip_name, 'w') as zipf:
        zipf.write(filepath, os.path.basename(filepath))
        return filepath

def create_xml(sheet1, sheet2, config, now_local):
    thisUID = str(uuid.uuid4())
    reportdate = current_datetime(parse_date_creation(str(sheet1["A5"].value))) #str(now_local)
    quarter , year = parse_quarter_year(sheet1["A3"].value)
    contractdate,gozuid = parse_contract_info(str(sheet1["A4"].value))
    root = ET.Element("Message")
    root.set('CreateDate', str(now_local))
    root.set('UID', str(thisUID))
    root.set('PreviousUID', str(config["PreviousUID"]))
    config["PreviousUID"] = thisUID
    Organization = ET.SubElement(root, "Organization")
    Organization.set("INN", str(config["INN"]))
    Organization.set("KPP", config["KPP"])
    Organization.set('Name', config["NAME"])
    Organization.set("GOZUID", str(gozuid))
    Organization.set("ContractDate", str(current_datetime(contractdate)))
    Forms = ET.SubElement(root, "Forms")
    Cash = ET.SubElement(Forms, "Cash",
                         {"ReportDate": reportdate, 'Year': year,'Quarter': quarter})
    contract_spending_mapping = {
        '1.': "Total",
        '1.1': "Salary",
        '1.2': "Taxes",
        '1.3': "Rates",
        '1.4': "OtherTotal",
        '1.4.1': "Return",
        '1.4.2': "Repayment",
        '1.4.3': "Another",
        '1.5': "Reserve",
        '1.6': "Income",
    }

    contract_finance_mapping = {
        '2.': "PlannedPay",
        '3.': "TotalRequirement",
        '4.': "CashBalance",
        '5.': "PlannedIncome",
        '6.': "PastPayments",
        '6.1': "SeparateAccount",
        '6.2': "BankAccount"
    }
   
    sorted_spending_keys = sorted(contract_spending_mapping, key=len, reverse=True)
    sorted_finance_keys = sorted(contract_finance_mapping, key=len, reverse=True)
    ContractSpending = ET.SubElement(Cash, "ContractSpending")
    ContractFinance = ET.SubElement(Cash, "ContractFinance")
    for row in range(11, sheet1.max_row + 1):
        
     cell_key_value = sheet1.cell(row=row, column=1).value
     cell_amount_value = sheet1.cell(row=row, column=3).value
     if cell_key_value is None:
      continue
     cell_key = str(cell_key_value)
    
     cell_amount = int(float(cell_amount_value.replace(',','.')) * 100) if cell_amount_value is not None else None
    
     if cell_amount is not None:
       spending_key = next((k for k in sorted_spending_keys if cell_key==k), None)
       if spending_key:
         
            ContractSpending.set(contract_spending_mapping[spending_key], str(cell_amount))
       finance_key = next((k for k in sorted_finance_keys if cell_key==k), None)
       if finance_key:
            ContractFinance.set(contract_finance_mapping[finance_key], str(cell_amount))

    FormSupplement = ET.SubElement(Forms, 'Supplement', {'ReportDate': reportdate})
    Parts = ET.SubElement(FormSupplement, "Parts")
    for row in range(3, sheet1.max_row + 1):
        if str(sheet2.cell(row=row, column=1).value).startswith('202'):
            Part = ET.SubElement(Parts, "Part", {"Year": str(sheet2.cell(row=row, column=1).value),
                                                 "Quarter": str(sheet2.cell(row=row, column=2).value),
                                                 "Requirement": str(int(sheet2.cell(row=row, column=3).value * 100)),
                                                 "Deviation": str(int(sheet2.cell(row=row, column=4).value * 100))})

            reasons = str(sheet2.cell(row=row, column=5).value).split(',')

            Reasons = ET.SubElement(Part, "Reasons")
            for reason in reasons:
                if reason.strip() != 'None':  # and reason.strip() != '':
                    ET.SubElement(Reasons, "Reason").text = reason.strip()

    from xml.dom.minidom import parseString
    xmlstr = parseString(ET.tostring(root, encoding="UTF-8", xml_declaration=True)).toprettyxml(encoding="UTF-8")

    return xmlstr, config


def update_config(filename, config):
    with open(filename, "w", encoding='utf-8') as configfile:
        config_json = json.dumps(config, ensure_ascii=False)
        configfile.write(config_json)


def save_xml(xmlstr, filename):
    with open(filename, "wb") as f:
        f.write(xmlstr)


def main():
    root = tk.Tk()
    root.withdraw()

    from tkinter import filedialog
    xlsx_file_path = filedialog.askopenfilename(title="Выберите XLSX файл", filetypes=[("XLSX Files", "*.xlsx")])
    config = load_config("config_v2.json")
    try:
        workbook = load_workbook_data(xlsx_file_path)
        sheet1 = workbook.worksheets[0]
        sheet2 = workbook.worksheets[1]
        xmlstr, updated_config = create_xml(sheet1, sheet2, config, current_datetime())
    except SheetTitleException as e:
        messagebox.ERROR(e)
    save_xml(xmlstr, os.path.dirname(xlsx_file_path) + "\message.xml")
    update_config("config_v2.json", updated_config)
    add_to_zip(os.path.dirname(xlsx_file_path) + '\message.xml', config["INN"], config["PreviousUID"])
    messagebox.showinfo("SAIBIS", "XML отчет сохранен рядом с " + xlsx_file_path)

    from lxml import etree
    try:
        with open('message_v5.xsd', 'rb') as f:
            schema_root = etree.XML(f.read())
        schema = etree.XMLSchema(schema_root)
        try:
            with open(os.path.dirname(xlsx_file_path) + '\message.xml', 'rb') as f:
                xml_root = etree.XML(f.read())
            try:
                schema.assertValid(xml_root)
                messagebox.showinfo("saibis", "XML соответствует XSD схеме.")
            except etree.DocumentInvalid as err:
                messagebox.showinfo("Ошипка", "Ошибка валидации:" + str(err))
        except IOError as err:
            print(err)
    except IOError as err:
        print(err)


if __name__ == "__main__":
    main()